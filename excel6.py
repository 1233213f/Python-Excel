import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter

# 打开excel文件
wb = openpyxl.load_workbook('data1.xlsx')
sheet = wb.active  # 正对表格

n1 = sheet.max_row
n2 = sheet.max_column

print(sheet.cell(10,1).value)
print(sheet.cell(11,1).value)

data = [[] for _ in range(n1)]
list = [[] for _ in range(n2)]

for column in range(1,n1+1):
    for row in range(1,n2+1):
        list[row-1] = sheet.cell(column, row).value
    data[column-1]=list
    list = [[] for _ in range(n2)]

# 在内存创建一个工作簿obj
wb = Workbook()
ws = wb.active
# ws.title = u'S'
# 向第一个sheet页写数据吧
i = 1
r = 1
for line in data:
    for col in range(1, len(line) + 1):
        ColNum = r
        ws.cell(row=r, column=col).value = line[col - 1]
    i += 1
    r += 1

wb.save('test1.xlsx')