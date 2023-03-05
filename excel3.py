import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# 咱们的表对象可以想列表或者字符串那样进行切片操作，
# 来获取电子表格中一行、一列或一个矩形区域中的所有 Cell 对象。
# 具体的用法如下： - 表对象[位置1:位置2] - 获取指定范围中的所有的单元格

# 打开excel文件
wb = openpyxl.load_workbook('GRB射线数据记录——后续补充.xlsx')

# 获取表
sheet = wb.active  # 正对表格

# 1.获取整个一行的单元格
max_column = sheet.max_column  # 获取最大列数
column = get_column_letter(max_column)  # 获取最大列数对应的字母列号
# 获取第一行所有单元格对象
row2 = sheet['A1':'%s1' % column]  # ((<Cell '表1'.A1>, <Cell '表1'.B1>, <Cell '表1'.C1>),)

for row_cells in row2:
    for cell in row_cells:
        print(cell.coordinate, cell.value)

# 2.获取整个列的单元格
max_row = sheet.max_row
columnB = sheet['B1':'B%d' % max_row]  # 获取B列对应的所有单元格对象
# 获取B列对应的所有单元格对象
for column_cells in columnB:
    for cell in column_cells:
        print(cell.coordinate, cell.value)

# 3. 获取矩形区域中的单元格对象
cell_tuples = sheet['A1': 'C3']
for cells in cell_tuples:
    for cell in cells:
        print(cell.coordinate, cell.value)