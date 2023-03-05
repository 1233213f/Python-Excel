import openpyxl

# 打开excel文件
wb = openpyxl.load_workbook('GRB射线数据记录——后续补充.xlsx')

# 获取活跃表对象
sheet = wb.active

# 获取单元格对应的 Cell 对象
a1 = sheet['A1']  # A1 表示A列中的第一行，这儿的列号采用的是从A开始的
print(a1)

# 获取单元格中的内容
content = a1.value
print(content)  # 结果是: a1中的内容

# 获取单元格的行和列信息
row = a1.row
print('行:', row)  # 结果： 单元格的行

column = a1.column
print('列:', column)  # 结果： 单元格的列

coordinate = a1.coordinate
print(coordinate)  # 结果：单元格的位置

# 获取第二列的所有内容
row_num = sheet.max_row  # 获取当前表中最大的行数
print(row_num)
column_num = sheet.max_column # 获取当前表中最大的列数
print(column_num)
#
# for row in range(1, row_num + 1):
#     cell1 = sheet.cell(row, 4)
#     print(cell1.value)

for row in range(1, column_num + 1):
    cell2 = sheet.cell(1, row)
    print(cell2.value)