# coding:utf-8
from openpyxl import load_workbook
import openpyxl
# 写入已存在的xlsx文件第一种方法
class Write_excel(object):
    '''修改excel数据'''
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet

    def write(self, row_n, col_n, value):
        '''写入数据，如(2,3,"hello"),第二行第三列写入数据"hello"'''
        self.ws.cell(row_n, col_n,value )
        self.wb.save(self.filename)

we = Write_excel("1.xlsx")
we.write(2,2,23)

# #写入已存在的xlsx文件第二种方法
# wb = load_workbook("1.xlsx")#生成一个已存在的wookbook对象
# wb1 = wb.active#激活sheet
# wb1.cell(2,2,'pass2')#往sheet中的第二行第二列写入‘pass2’的数据
# wb.save("1.xlsx")#保存